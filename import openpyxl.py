import openpyxl
import os
import re
import datetime

# 打開 Excel 文件
workbook = openpyxl.load_workbook('IQ-GMB 資料庫資料表說明 (3).xlsx')
sheet = workbook.active

# 檢查文件是否存在
file_path = "output_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + ".sql"
# 全新覆寫檔案
with open(file_path, "w") as file:
    file.write("")

two_null_encountered = False
started_printing = False
table_names = []
table_name = ""
# 遍歷工作表的每一行，依照條件打印每一行的值
for row in sheet.iter_rows(values_only=True):
    if row[0] and " - " in row[0]:
        # 使用" - "分割文字，並取第一部分作(split(" - ")[0])為表格名稱
        table_name = row[0].split(" - ")[0]
        table_names.append(table_name)

    # 去掉Excel一開始的大標題，設定開始遍歷的位置
    if not started_printing:
        if row[0] is not None and " - " in row[0]:
            started_printing = True
            #print(row)
    else:
        if row[0] is None:
            if two_null_encountered:
                break
            else:
                two_null_encountered = True
        else:
            two_null_encountered = False
            if " - " in row[0]:
                two_null_encountered = False
            # 欄位名稱只能是 1.英文大小寫 2.數字 3.底線
            if re.sub(r'[^a-zA-Z0-9_]', '', row[0]) == row[0]:
                for i in range(6):  # row 有 6 個元素(欄位)，從 row[0] 到 row[5]
                    row_list = list(row) # 將 tuple 轉換為 list
                    if row_list[i] is not None: # 值不為None的情況下，將單引號改為雙引號
                        row_list[i] = row_list[i].replace("'", '"')
                row = tuple(row_list)  # 將 list 轉換回 tuple
                sql_code = f"""if exists (SELECT * FROM INFORMATION_SCHEMA.COLUMNS where TABLE_NAME = '{table_name}' AND COLUMN_NAME = '{row[0]}')
BEGIN             
    IF NOT EXISTS (SELECT 1 FROM fn_listextendedproperty(N'MS_Description', 'SCHEMA', 'dbo', 'TABLE', '{table_name}', 'COLUMN', '{row[0]}'))
    BEGIN
        -- 新增欄位備註
        EXEC sys.sp_addextendedproperty @name=N'MS_Description', @value=N'{row[5]}' , @level0type=N'SCHEMA',@level0name=N'dbo', @level1type=N'TABLE',@level1name=N'{table_name}', @level2type=N'COLUMN',@level2name=N'{row[0]}'
    END
    else
    BEGIN
        -- 更新欄位備註
        EXEC sys.sp_updateextendedproperty @name=N'MS_Description', @value=N'{row[5]}' , @level0type=N'SCHEMA',@level0name=N'dbo', @level1type=N'TABLE',@level1name=N'{table_name}', @level2type=N'COLUMN',@level2name=N'{row[0]}'
    END
END
else
BEGIN
    print('Error Table: ' + '{table_name}')
END
"""
            else:
                continue
            
            with open(file_path, "a", encoding='UTF-8') as file:
                file.write("\n\n")
                file.write(sql_code)
# 提取包含" - "的字段，並只取" - "符號前面的值
# TableName = [row[0].split(" - ")[0] for row in sheet.iter_rows(values_only=True) if row[0] and " - " in row[0]]
# print(table_names)
# print("資料表總數: " + str(len(table_names)))




# 關閉 Excel 文件
workbook.close()
